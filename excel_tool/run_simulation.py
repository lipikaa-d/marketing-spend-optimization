import pandas as pd
import numpy as np
import joblib
from openpyxl import load_workbook

# File paths
EXCEL_PATH = "excel_tool/Marketing_Spend_Simulator.xlsx"
MODEL_PATH = "models/final_log_linear_model.pkl"
FEATURES_PATH = "models/model_features.pkl"

# Load model and features
model = joblib.load(MODEL_PATH)
features = joblib.load(FEATURES_PATH)


# Read Excel inputs

df_inputs = pd.read_excel(EXCEL_PATH, sheet_name="Inputs")

# Create dictionary: Channel -> Value
input_dict = dict(zip(df_inputs["Channel"], df_inputs["Value"]))

uplift_pct = input_dict.pop("Uplift_Percentage") / 100.0

# Convert to DataFrame in correct feature order
base_input = pd.DataFrame([[input_dict[f] for f in features]], columns=features)


# Baseline prediction
log_base_sales = model.predict(base_input)[0]
base_sales = np.expm1(log_base_sales)


# Scenario prediction
scenario_input = base_input * (1 + uplift_pct)
log_scenario_sales = model.predict(scenario_input)[0]
scenario_sales = np.expm1(log_scenario_sales)


# Metrics
absolute_lift = scenario_sales - base_sales
percent_lift = (absolute_lift / base_sales) * 100


# Write outputs to Excel

wb = load_workbook(EXCEL_PATH)
ws = wb["Outputs"]

ws["B2"] = round(float(base_sales), 2)
ws["B3"] = round(float(scenario_sales), 2)
ws["B4"] = round(float(absolute_lift), 2)
ws["B5"] = round(float(percent_lift), 2)

wb.save(EXCEL_PATH)

print("Simulation completed successfully.")
